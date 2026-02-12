from faker import Faker
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
fake_data = Faker(['hi_IN', 'en_US'])

for i in range(1, 11):  # row
    for j in range(1, 5):  #column
        ws.cell(row=i, column=1).value = fake_data.name()
        ws.cell(row=i, column=2).value = fake_data.email()
        ws.cell(row=i, column=3).value = fake_data.phone_number()
        ws.cell(row=i, column=4).value = fake_data.address()
wb.save("testData.xlsx")

#print in one or single source of fake info record
'''print(fake_data.name())
print(fake_data.email())
print(fake_data.address())
print(fake_data.phone_number())'''

# now to generate 1k+ record